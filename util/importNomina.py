import time

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

if __name__ == '__main__':

    """ FUNCION SIN READONLY """
    time_start = time.time()
    wb = load_workbook(filename='nominaEXT.xlsx', data_only=True)  # , read_only=True)
    time_end = time.time()
    print(f"TIEMPO wb: {time_end - time_start}")

    ws = wb['NOMINA']
    dataE = coordinate_from_string('B6')
    nombres = {}

    time_start = time.time()
    for i, row in enumerate(ws.iter_rows(min_row=dataE[1], max_row=len(ws[dataE[0]]), min_col=column_index_from_string(dataE[0]),
                                         max_col=column_index_from_string(dataE[0]))):
        for cell in row:
            nombres[cell.value] = cell.offset(column=1).value
            print(f'{i}: {cell.value} - {cell.offset(column=1).value}')
    time_end = time.time()
    print(f"TIEMPO FOR1: {time_end - time_start}")

    """ FUNCION READONLY """
    # time_start = time.time()
    # wb = load_workbook(filename='nominaEXT.xlsx', data_only=True)  # , read_only=True)
    # time_end = time.time()
    # print(f"TIEMPO wb: {time_end - time_start}")
    #
    # ws = wb['NOMINA']
    # dataE = coordinate_from_string('B6')
    # nombres = {}
    #
    # time_start = time.time()
    # for row in ws.iter_rows(min_row=6, max_row=161, min_col=2, max_col=2):
    #     for cell in row:
    #         nombres[cell.value] = cell.value
    #
    # time_end = time.time()
    # print(f"TIEMPO FOR2: {time_end - time_start}")
