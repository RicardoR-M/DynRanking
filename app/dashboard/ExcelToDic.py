import warnings
from pprint import pprint
from time import strptime

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def load_nomina(obj):
    wb = load_workbook(filename=obj, data_only=True, keep_vba=False, keep_links=False)

    if "NOMINA" not in wb.sheetnames:
        raise RuntimeError('No se cargo el excel correcto, no se encontro la hoja NOMINA')

    ws = wb['NOMINA']
    cell_e = coordinate_from_string('B6')
    col_nombre = 'C'
    col_super = 'AB'
    dic_nomina = {}

    for i, row in enumerate(
            ws.iter_rows(min_row=cell_e[1],
                         max_row=len(ws[cell_e[0]]),
                         min_col=column_index_from_string(cell_e[0]),
                         max_col=column_index_from_string(cell_e[0]))):
        for cell in row:
            if cell.value is None:
                continue

            dic_nomina[cell.value] = {
                'e': cell.value,
                'nombre': str(ws[f'{col_nombre}{cell.row}'].value).strip(),
                'supervisor': str(ws[f'{col_super}{cell.row}'].value).strip()
            }
    print(dic_nomina)
    print(len(dic_nomina))
    return list(dic_nomina.values())


def load_cruzado(obj):
    warnings.simplefilter("ignore")  # Desactiva el warning de graficos cuando se utiliza load_workbook
    wb = load_workbook(filename=obj, data_only=True, keep_vba=False, keep_links=False)
    warnings.simplefilter("default")  # Habilita los warnings nuevamente

    if "DATA" not in wb.sheetnames:
        raise RuntimeError('No se cargo el excel correcto, no se encontro la hoja DATA')

    # Ubicación de la columna codigo E, se debe poner el primer dato, no se debe direccionar a la cabecera
    ws = wb["DATA"]
    cord_e = coordinate_from_string("D6")  # (D, 6)
    col_nota = "BR"
    col_proveedor = "H"
    proveedor = "DYNAMICALL"
    cruzado_list = []
    sn_offset = 7
    mes_offset = 8
    motivo_offset = 15
    fcr_offset = 73
    detalle_offset = 16
    observacion_offset = 81
    calificacion_offset = 66
    items_b1_offset = list(range(26, 31))
    items_b2_offset = list(range(32, 36))
    items_b3_offset = list(range(37, 48))
    items_b4_offset = list(range(49, 53))
    items_b5_offset = list(range(54, 57))
    items_b6_offset = 58

    for i, row in enumerate(ws.iter_rows(min_row=cord_e[1],
                                         max_row=len(ws[cord_e[0]]),
                                         min_col=column_index_from_string(cord_e[0]),
                                         max_col=column_index_from_string(cord_e[0]))):
        for cell in row:
            if cell.value is None:
                continue

            if ws[f"{col_proveedor}{cell.row}"].value != proveedor:
                continue

            # if cell.value in dic_cruzado:
            #     dic_cruzado[cell.value]["sumatoria"] = round(
            #         dic_cruzado[cell.value]["sumatoria"] + ws[f"{col_nota}{cell.row}"].value, 2)
            #     dic_cruzado[cell.value]["q"] += 1
            # else:
            #     dic_cruzado[cell.value] = {
            #         "e": str(cell.value).strip(),
            #         "sumatoria": round(ws[f"{col_nota}{cell.row}"].value, 2),
            #         "q": 1,
            #     }
            cruzado_list.append({
                'SN': cell.offset(column=sn_offset).value,
                'MesMonitoreo': get_month_name(strptime(str(cell.offset(column=mes_offset).value), '%Y-%m-%d %H:%M:%S').tm_mon),
                'TipoEvaluacion': 'Cruzado',
                'Asesor': str(cell.value).strip(),
                'Cbo_estado': cell.offset(column=motivo_offset).value,
                'Nivel_1': ('SI' if cell.offset(column=fcr_offset).value == '' else cell.offset(column=fcr_offset).value),
                'DetalleLlamada': cell.offset(column=detalle_offset).value,
                'ObservacionesAspectoMejora': cell.offset(column=observacion_offset).value,
                'Calificacion': round(cell.offset(column=calificacion_offset).value, 2),
                # bloque 1
                'SaludaCliente': cell.offset(column=items_b1_offset[0]).value,
                'SeDirigePorNombre': cell.offset(column=items_b1_offset[1]).value,
                'InteractuaConCliente': cell.offset(column=items_b1_offset[2]).value,
                'EvitaUsoTecnicismos': cell.offset(column=items_b1_offset[3]).value,
                'SeDespideComoIndicaManual': cell.offset(column=items_b1_offset[4]).value,
                # bloque 2
                'ValidaConsultaTransaccion': cell.offset(column=items_b2_offset[0]).value,
                'RealizaPreguntasPrecision': cell.offset(column=items_b2_offset[1]).value,
                'ValidaMotivoReal': cell.offset(column=items_b2_offset[2]).value,
                'ValidaAtencionPrevia': cell.offset(column=items_b2_offset[3]).value,
                # bloque 3
                'ValidaCES': cell.offset(column=items_b3_offset[0]).value,
                'AtencionPasoAPaso': cell.offset(column=items_b3_offset[1]).value,
                'SolicitaDNIRUC': cell.offset(column=items_b3_offset[2]).value,
                'ValidaTRACER': cell.offset(column=items_b3_offset[3]).value,
                'VerificaWebAverias': cell.offset(column=items_b3_offset[4]).value,
                'VerificaParametrosServicios': cell.offset(column=items_b3_offset[5]).value,
                'ProblemasInternetIngresaEMTA': cell.offset(column=items_b3_offset[6]).value,
                'ProblemasTelefoniaIngresaJANUS': cell.offset(column=items_b3_offset[7]).value,
                'ProblemasInternetSmarTV': cell.offset(column=items_b3_offset[8]).value,
                'ExplicacionBrindadaCorresponde': cell.offset(column=items_b3_offset[9]).value,
                'ValidaConClienteInformacion': cell.offset(column=items_b3_offset[10]).value,
                # bloque 4
                'EjecutaAccionesAplicativos': cell.offset(column=items_b4_offset[0]).value,
                'SeTipificaSIACSGA': cell.offset(column=items_b4_offset[1]).value,
                'NotasPlantillaTipificacion': cell.offset(column=items_b4_offset[2]).value,
                'SeTipificaEnSIAC': cell.offset(column=items_b4_offset[3]).value,
                # bloque 5
                'EvitaComentariosNegativos': cell.offset(column=items_b5_offset[0]).value,
                'EvitaPalabraSoeces': cell.offset(column=items_b5_offset[1]).value,
                'EscuchaClienteSinInterrumpirlo': cell.offset(column=items_b5_offset[2]).value,
                # bloque 6
                'MotivoRealNecesidad': cell.offset(column=items_b6_offset).value
            })
    return cruzado_list


def get_month_name(month_n):
    mes = {
        1: 'Enero',
        2: 'Febrero',
        3: 'Marzo',
        4: 'Abril',
        5: 'Mayo',
        6: 'Junio',
        7: 'Julio',
        8: 'Agosto',
        9: 'Septiembre',
        10: 'Octubre',
        11: 'Noviembre',
        12: 'Diciembre'
    }
    return mes[month_n]


if __name__ == '__main__':
    load_cruzado('cruzado.xlsx')
