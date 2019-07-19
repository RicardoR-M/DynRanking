import time
import warnings

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from app.dbase.dbLogic import BaseRanking

if __name__ == "__main__":
    time_start = time.time()
    cruzadoDB = BaseRanking()

    # Ubicación de la columna codigo E, se debe poner el primer dato, no se debe direccionar a la cabecera
    cordE = coordinate_from_string("D6")  # (D, 6)
    colNota = "BR"
    colProveedor = "H"
    proveedor = "DYNAMICALL"
    dicAsesores = {}

    """ FUNCION SIN READONLY """

    warnings.simplefilter("ignore")  # Desactiva el warning de graficos cuando se utiliza load_workbook
    wb = load_workbook(filename="cruzadoEXT.xlsx", data_only=True, keep_vba=False, keep_links=False)
    warnings.simplefilter("default")  # Habilita los warnings nuevamente

    if "DATA" not in wb.sheetnames:
        print("No se encontro la hoja DATA")
        quit()

    ws = wb["DATA"]

    for i, row in enumerate(ws.iter_rows(min_row=cordE[1], max_row=len(ws[cordE[0]]), min_col=column_index_from_string(cordE[0]),
                                         max_col=column_index_from_string(cordE[0]))):
        for cell in row:
            if cell.value is None:
                continue

            if ws[f"{colProveedor}{cell.row}"].value != proveedor:
                continue

            if cell.value in dicAsesores:
                dicAsesores[cell.value]["sumatoria"] = round(
                    dicAsesores[cell.value]["sumatoria"] + ws[f"{colNota}{cell.row}"].value, 2)
                dicAsesores[cell.value]["q"] += 1
            else:
                dicAsesores[cell.value] = {
                    "e": cell.value,
                    "sumatoria": ws[f"{colNota}{cell.row}"].value,
                    "q": 1,
                }

    time_end = time.time()
    # print(dicAsesores.get("MAX", "No se encontro asesor"))

    cruzadoDB.insert_cruzado(list(dicAsesores.values()))

    print(f"TIEMPO TOTAL: {time_end - time_start} - MAX ROW: {len(ws[cordE[0]])} - DIC SIZE: {len(dicAsesores)}")
