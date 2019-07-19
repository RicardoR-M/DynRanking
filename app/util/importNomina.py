import time
from pprint import pprint

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from app.dbase.dbLogic import BaseRanking

if __name__ == '__main__':
    db = BaseRanking()

    """ FUNCION SIN READONLY """
    time_startload = time.time()
    wb = load_workbook(filename='nominaEXT.xlsx', data_only=True, keep_vba=False, keep_links=False)  # , read_only=True)
    time_endload = time.time()

    if "NOMINA" not in wb.sheetnames:
        print("No se encontro la hoja NOMINA")
        quit()

    ws = wb['NOMINA']
    cellE = coordinate_from_string('B6')
    colNombre = 'C'
    colSuper = 'N'
    nomina = {}

    time_start = time.time()
    for i, row in enumerate(ws.iter_rows(min_row=cellE[1], max_row=len(ws[cellE[0]]), min_col=column_index_from_string(cellE[0]),
                                         max_col=column_index_from_string(cellE[0]))):
        for cell in row:
            if cell.value is None:
                continue

            nomina[cell.value] = {
                'e': cell.value,
                'nombre': ws[f'{colNombre}{cell.row}'].value,
                'supervisor': ws[f'{colSuper}{cell.row}'].value
            }

    time_sql = time.time()
    db.insert_nomina(list(nomina.values()))
    time_sqle = time.time()

    pprint(nomina)
    print(len(nomina))
    time_end = time.time()
    print(f"TIEMPO SQLInsert: {time_sqle - time_sql}")
    print(f"TIEMPO Load wb: {time_endload - time_startload}")
    print(f"TIEMPO Lectura wb: {time_end - time_start}")
